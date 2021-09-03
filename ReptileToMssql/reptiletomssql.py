# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import glob

import xlrd  # xlrd操作Excel表
# import openpyxl  # openpyxl操作Excel表
from openpyxl import Workbook  # openpyxl操作Excel表
from openpyxl import load_workbook  # openpyxl操作Excel表
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as pq
import time  # 延时或者控件使用
import pymssql  # 链接Mssql
import os
import sys  # 系统自带

# import os  # os 插件

url = 'https://fanyi.baidu.com/?aldtype=16047#auto/zh'  # 在线翻译网址
driver = webdriver.Chrome()
wait = WebDriverWait(driver, 10)
conn = pymssql.connect(server='DESKTOP-CG7QTRM\\TESTMSSQLSERVER', user='sa', password='123', database='DBServerTest')

commlist = []


# 等待页面加载完成
def search():
    print('正在检索')
    try:
        # 等待页面全部加载完毕
        wait.until(
            # 隐式等待，直接下面指定的元素可见
            EC.presence_of_element_located(
                (By.CSS_SELECTOR,
                 '.container .main .main .inner .translate-wrap .translateio .translate-main .trans-right .output-wrap .output-mod .output-bd .ordinary-output'))
        )
        html = driver.page_source  # 返回页面源码
        return html
    except TimeoutException:  # 超时异常
        return search()

english=[]
Num=1

# 这个可以优化修改,可以用迭代
def check_english_distinct(englishtmp):
    if(englishtmp in english ):
        englishtmp = englishtmp +Num
        check_english_distinct(englishtmp)
        Num= Num+1
        return englishtmp


# 返回的html去找指定的text数据
def parse_one_page(html, name):
    doc = pq(html)
    items = doc(
        '.container .main .main .inner .translate-wrap .translateio .translate-main .trans-right .output-wrap').items()
    for item in items:
        englishtmp='c' +item.find('.output-mod .output-bd .ordinary-output').text().replace(name, "").replace(" ","")
        if(englishtmp=='c'):
            englishtmp='c' +name

        comments = {
            "English": englishtmp,
            "Chinese": name
        }
        Num=1
        if(englishtmp=='cSerialnumber'):
           tt='ss'
        englishtmp=check_english_distinct(englishtmp)


        english.append(englishtmp)

        commlist.append(comments)
        print(comments)
        # save_to_mongo(comments)
    # next_page()
    time.sleep(1)


# 翻页操作
# def next_page():
#     try:
#        '#tie-main > div.tie-new > div.list-foot.clearfix > div > ul > li:nth-child(6) > span'
#         '//*[@id="tie-main"]/div[3]/div[3]/div/ul/li[6]/span'
#         if wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,
#                                                      '.wrapper .main-bg.clearfix #tie-main .tie-new .list-foot.clearfix .page-bar .m-page .next.z-enable'))):
#             next_page = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,
#                                                                    '.wrapper .main-bg.clearfix #tie-main .tie-new .list-foot.clearfix .page-bar .m-page .next.z-enable')))
#             next_page.click()
#     except TimeoutException:
#         return None

# 保存到Excel
def read_excel_openpyxl(path):
    wb = load_workbook(path)
    # wb = load_workbook(r"C:\Users\Twj\Desktop\新人电脑配件.xlsx")
    # wb = load_workbook(r"C:\Users\Twj\Desktop\新人电脑配件.xlsx")
    ws = wb["sheet1"]
    for rows in ws.iter_rows(min_row=1, max_row=20, min_col=1):
        for cell in rows:
            print('cell %s %s' % (cell.coordinate, cell.value))
    print(wb.sheetnames)

#xlrd的读取
def read_excel_xlrd(path):
    # 打开文件
    # workbook = xlrd.open_workbook(path)
    workbook = xlrd.open_workbook(r'C:\Users\Twj\Desktop\新人电脑配件.xlsx')
    # workbook = xlrd.open_workbook(r'C:\Users\Twj\Desktop\客户PCN变更1.xlsx')
    # 获取所有sheet
    names = workbook.sheet_names()
    print(workbook.sheet_names())
    # 方法①
    ret1 = workbook.sheets()[0]
    # 方法②
    # ret2 = excel_content.sheet_by_index(0)
    # 方法③
    # ret 3 = excel_content.sheet_by_name("动物类别")


# table = workbook.sheets()[0]

# 这个可以使用
# def exec_sql_insert():
# sql = r"Insert into CC_tChineseToEnglish (Chinese,English) values ('" + chinese + "','" + endlish + "')"
# insert_mssql(sql)
# cursor_1 = conn.cursor(as_dict=True)
# cursor_1.executemany('Insert into CC_tChineseToEnglish (Chinese,English) values (%s,%s)', [(chinese, endlish)])
# conn.commit()


# def create_txt_file(file_name):
    #  判断文件是否存在，不存在就创建
    # folder = os.path.exists("F:\PythonCreateTableLog\\" + file_name + '.txt')
    # if not folder:  # 判断是否存在文件夹如果不存在则创建为文件夹
    #     os.makedirs("F:\PythonCreateTableLog\\" + file_name + '.txt')  # makedirs 创建文件时如果路径不存在会创建这个路径


# 保存日志到本地log
def save_txt_file(file_name, string):
    # 保存到本地
    # path="F:\PythonCreateTableLog\\" + file_name + '.txt'
    f = open("F:\PythonCreateTableLog\\" + file_name + '.txt', "a")  # w是覆盖,a是追加
    f.write(string)  # 这句话自带文件关闭功能，不需要再写f.close()
    f.close()


# 生成SQL代码创建
def exec_sql_crate_table(table_name):
    stringSql = "SELECT distinct TableName FROM CC_tTmp "
    list = find_mssql(stringSql)
    for listtmp in list:   
        create_sql = ''

        stringSql1 = "SELECT ColumnsName FROM CC_tTmp where TableName='{TableName}' order by orderby".format(
            TableName=listtmp[0])
        list2 = find_mssql(stringSql1) 

        # 爬虫 没返回，直接用全局变量
        get_pachong(list2)
      
        # 组合列
        for item in commlist:
            create_sql += ',[{columnName}] varchar(100)'.format(columnName=item.get('English',''))
        create_sql += ',cupdatauser varchar(100),cupdatatime datetime'
        create_sql = 'create table ' + listtmp[0] + ' ( cID int primary key identity(1,1)' + create_sql + ')'
        cursor_1 = conn.cursor(as_dict=True)
        cursor_1.execute(create_sql)
        conn.commit()
        # 保存SQL代码
        save_txt_file(listtmp[0]+'Table', create_sql+'\n')

        # 创建说明，中英文对照表
        for item in commlist:
            sql = "execute sp_addextendedproperty 'MS_Description', '{Chinese}', 'SCHEMA', 'dbo', 'table', '{tablename}', 'column', '{English}'" \
                .format(English=item.get('English', ''), Chinese=item.get('Chinese', ''),tablename=listtmp[0])
            cursor_1.execute(sql)
            conn.commit()
            # 保存中英文对照表
            save_txt_file(listtmp[0]+'property', sql+'\n')
        sql = "execute sp_addextendedproperty 'MS_Description', '修改人', 'SCHEMA', 'dbo', 'table', '{tablename}', 'column', 'cupdatauser'".format(tablename=listtmp[0])
        cursor_1.execute(sql)
        conn.commit()
        sql = "execute sp_addextendedproperty 'MS_Description', '主键ID', 'SCHEMA', 'dbo', 'table', '{tablename}', 'column', 'ID'".format(tablename=listtmp[0])
        cursor_1.execute(sql)
        conn.commit()
        sql = "execute sp_addextendedproperty 'MS_Description', '修改时间', 'SCHEMA', 'dbo', 'table', '{tablename}', 'column', 'cupdatatime'".format(tablename=listtmp[0])
        cursor_1.execute(sql)
        conn.commit()
        english.clear()
        commlist.clear()


def get_pachong(list2):
    for tmp in list2:
        driver.get(url + '/' + tmp[0].split('/')[0])
        time.sleep(2)
        html = search()
        parse_one_page(html, tmp[0].split('/')[0])


def exec_sql_find():
    sql = "SELECT * FROM CC_tGetNumber"
    find_mssql(sql)


def getpath_files():
    path = 'C:\\Users\\Twj\\Desktop\\Path'
    # f2 = open(f, 'a+')
    for filename in os.listdir(path):
        if filename.endswith('.xlsx'):
            read_excel_xlrd(path + '\\' + filename)


# 数据库连接查询
def find_mssql(sql):
    cursor_1 = conn.cursor()
    cursor_1.execute(sql)
    # print(cursor_1.fetchall())
    return cursor_1.fetchall()


# 数据库链接插入
def insert_mssql(sql):
    cursor_1 = conn.cursor()
    cursor_1.executemany(sql)
    conn.commit()


def main():
    # exec_sql_find()

    # read_excel('')
    # find_mssql()
    # listtmp = getpath_files()

    # for value in listtmp:
    #     driver.get(url + '/' + value)
    #     time.sleep(2)
    #     html = search()
    #     parse_one_page(html, value)
    # driver.close()
    exec_sql_crate_table('')
    driver.close()


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()
else:
    print('false')
# https://fanyi.baidu.com/?aldtype=16047#auto/zh
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
